import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):

      wb = xl.load_workbook('transactions.xlsx')
      sheet = wb['Sheet1']

      # Calculate corrected prices
      for row in range(2, sheet.max_row + 1):
          original_price = sheet.cell(row=row, column=3).value

          if original_price is not None:
              corrected_price = original_price * 0.9
              sheet.cell(row=row, column=4).value = corrected_price

      # Create chart AFTER loop
      values = Reference(
          sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4
      )

      chart = BarChart()
      chart.add_data(values)
      sheet.add_chart(chart, "E2")

      wb.save(filename)
      print("Chart added successfully!")
